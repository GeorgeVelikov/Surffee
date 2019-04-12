from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views.generic import CreateView, UpdateView
from django_countries.fields import Country

from surveys.models import SurveyAnswer
from ..models.survey import Survey, Choice, Question
from ..forms.surveys import ResearcherCreateSurvey
from ..forms.users import ResearcherCreationForm

from .error import permission_user_logged_in, permission_user_owns_survey

from django.core.exceptions import PermissionDenied
from ast import literal_eval

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color


class Create(UpdateView):
    template_name = 'surveys/create_survey.html'
    model = Survey
    form_class = ResearcherCreateSurvey

    # can just increment id of the last survey
    def get(self, request, *args, **kwargs):
        permission_user_logged_in(request)

        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        return self.render_to_response(self.get_context_data(form=form))

    def post(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)

        request.POST._mutable = True
        form.data['creator'] = request.user.pk
        # currently saves the PI we want in the shape of a list which is translated to a string, fix this later w/ eval
        form.data['pi_choices'] = request.POST.getlist('pi_set')
        request.POST._mutable = False

        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        self.object = form.save(commit=True)
        return redirect('/surveys/inactive')

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))


class ActiveToggle(UpdateView):
    model = Survey

    def get(self, request, *args, **kwargs):
        survey_id = self.kwargs.get('survey_id')
        survey = Survey.objects.get(pk=survey_id)

        permission_user_logged_in(request)
        permission_user_owns_survey(request, survey)

        if survey.active:
            survey.active = False
        else:
            survey.active = True

        survey.save()
        return redirect('/surveys/' + str(survey_id))


class Delete(UpdateView):
    model = Survey

    def get(self, request, *args, **kwargs):
        survey_id = self.kwargs.get('survey_id')
        survey = Survey.objects.get(pk=survey_id)

        permission_user_logged_in(request)
        permission_user_owns_survey(request, survey)

        questions = Question.objects.filter(survey=survey)
        choices = Choice.objects.filter(question__in=questions)

        choices.delete()
        questions.delete()
        survey.delete()
        return redirect('/surveys/')


def detail(request, survey_id):
    if not request.user.is_authenticated:  # user is not logged in
        raise PermissionDenied("User is not logged in")

    template = 'surveys/detail.html'
    survey = Survey.objects.get(pk=survey_id)

    if survey.creator != request.user:
        raise PermissionDenied("You do not own this survey")

    # this converts the string representation of a list back to a list
    if survey.pi_choices:
        choices = literal_eval(survey.pi_choices)
        for i in range(len(choices)):
            choices[i] = choices[i].replace("_", " ").capitalize()
    else:
        choices = None

    if survey.creator != request.user and not request.user.is_superuser:
        raise PermissionDenied("You have tried to access " + survey.name + ". To gain permissions please contact "
                               + survey.creator.email + ".")
    context = {'survey': survey, 'choices': choices}
    return render(request, template, context)


def export(request, survey_id):
    def create_sheet(title, objects, columns, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title

        # headers
        create_headers(ws, columns)

        # content
        row = 2
        for object in objects:
            for i in range(len(columns)):
                attribute = eval("object." + columns[i])
                cell = ws.cell(row=row, column=i + 1)
                cell.value = str(attribute) if type(attribute) == Country else attribute
            row += 1

    def create_headers(sheet, columns):
        for col_num in range(len(columns)):
            cell = sheet.cell(row=1, column=col_num + 1)
            dot_index = 0 if columns[col_num].rfind('.') == -1 else columns[col_num].rfind('.')+1
            cell.value = columns[col_num][dot_index:].replace("_", " ").capitalize()
            cell.font = Font(bold=True)

    surveys = Survey.objects.filter(pk=survey_id)
    survey = surveys[0]
    filename = survey.name.replace(" ", "-") + '.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = openpyxl.Workbook()

    create_sheet("Survey", surveys, ["name", "creator.username", "description", "pi_choices"], first=True)

    questions = survey.question_set.all()
    create_sheet("Questions", questions, ["question_text", "type"])

    choices = Choice.objects.filter(question__in=questions)
    create_sheet("Choices", choices, ["question.question_text", "choice_text", "votes"])

    # Answer sheet
    answers_sheet = wb.create_sheet("Answers")
    answers = SurveyAnswer.objects.filter(survey=survey)
    pi_columns = ["ip_address"] + [("pi_questions." + pi) for pi in eval(survey.pi_choices)]
    answer_ids = []
    question_ids = []
    row = 2

    # create headers for personal info columns
    create_headers(answers_sheet, pi_columns)

    # create headers for every question in the survey
    questions_start_col = len(pi_columns) + 1
    col = questions_start_col
    for question in survey.question_set.all():
        cell = answers_sheet.cell(1, col)
        cell.value = question.question_text
        cell.font = Font(bold=True)
        question_ids.append(question.pk)
        col += 1

    # populate rows with personal info
    for answer in answers:
        for i in range(len(pi_columns)):
            attribute = eval("answer." + pi_columns[i])
            c = answers_sheet.cell(row=row, column=i + 1)
            c.value = str(attribute) if type(attribute) == Country else attribute
        answer_ids.append(answer.pk)
        row += 1

    # populate rows with selected choices for given questions
    for i in range(answers.count()):
        for a in range(col - questions_start_col):
            cell = answers_sheet.cell(i + 2, questions_start_col + a)
            choices = Choice.objects.filter(surveyanswer__pk=answer_ids[i], question__pk=question_ids[a])
            cell_content = ""
            for choice in choices:
                cell_content += (choice.choice_text + ", ")
            cell.value = str(cell_content[:-2])

    wb.save(response)
    return response
