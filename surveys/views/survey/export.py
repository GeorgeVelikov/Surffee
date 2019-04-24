from django.http import HttpResponse

from django_countries.fields import Country

from surveys.models import SurveyAnswer
from ...models.survey import Survey, Choice

import openpyxl
from openpyxl.styles import Font


def export(request, survey_id):
    def create_sheet(title, objects, columns, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title

        # headers
        create_headers(ws, columns)

        # content
        row = 2
        for object in objects:
            for i in range(len(columns)):
                attribute = eval("object." + columns[i])
                cell = ws.cell(row=row, column=i + 1)
                cell.value = str(attribute) if type(attribute) == Country else attribute
            row += 1

    def create_headers(sheet, columns):
        for col_num in range(len(columns)):
            cell = sheet.cell(row=1, column=col_num + 1)
            dot_index = 0 if columns[col_num].rfind('.') == -1 else columns[col_num].rfind('.')+1
            cell.value = columns[col_num][dot_index:].replace("_", " ").capitalize()
            cell.font = Font(bold=True)

    surveys = Survey.objects.filter(pk=survey_id)
    survey = surveys[0]
    filename = survey.name.replace(" ", "-") + '.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = openpyxl.Workbook()

    create_sheet("Survey", surveys, ["name", "creator.username", "description", "pi_choices"], first=True)

    questions = survey.question_set.all()
    create_sheet("Questions", questions, ["question_text", "type"])

    choices = Choice.objects.filter(question__in=questions)
    create_sheet("Choices", choices, ["question.question_text", "choice_text", "votes"])

    # Answer sheet
    answers_sheet = wb.create_sheet("Answers")
    answers = SurveyAnswer.objects.filter(survey=survey)
    pi_columns = ["ip_address"] + [("pi_questions." + pi) for pi in eval(survey.pi_choices)]
    answer_ids = []
    question_ids = []
    row = 2

    # create headers for personal info columns
    create_headers(answers_sheet, pi_columns)

    # create headers for every question in the survey
    questions_start_col = len(pi_columns) + 1
    col = questions_start_col
    for question in survey.question_set.all():
        cell = answers_sheet.cell(1, col)
        cell.value = question.question_text
        cell.font = Font(bold=True)
        question_ids.append(question.pk)
        col += 1

    # populate rows with personal info
    for answer in answers:
        for i in range(len(pi_columns)):
            attribute = eval("answer." + pi_columns[i])
            c = answers_sheet.cell(row=row, column=i + 1)
            c.value = str(attribute) if type(attribute) == Country else attribute
        answer_ids.append(answer.pk)
        row += 1

    # populate rows with selected choices for given questions
    for i in range(answers.count()):
        for a in range(col - questions_start_col):
            cell = answers_sheet.cell(i + 2, questions_start_col + a)
            choices = Choice.objects.filter(surveyanswer__pk=answer_ids[i], question__pk=question_ids[a])
            cell_content = ""
            for choice in choices:
                cell_content += (choice.choice_text + ", ")
            cell.value = str(cell_content[:-2])

    wb.save(response)
    return response
